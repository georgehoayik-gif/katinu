import requests
import pandas as pd
import streamlit as st
import plotly.express as px
import json
from datetime import datetime, timedelta
from PIL import Image, ImageDraw, ImageFont
import os
import time

# =========================================
# 💾 CACHE LOCAL SQLITE (MINI BANCO)
# ========================================= to_sql

import sqlite3
st.write("App iniciado com sucesso ??")
DB_PATH = "/tmp/cache_milvus.db"

# =========================================
# 🚀 CACHE POR PÁGINA (NÍVEL HARDCORE)
# =========================================
def carregar_pagina_cache(chave):
    if not os.path.exists(DB_PATH):
        return pd.DataFrame()

    conn = sqlite3.connect(DB_PATH)

    try:
        df = pd.read_sql(
            f"SELECT * FROM cache_paginas WHERE cache_key = ?",
            conn,
            params=(chave,)
        )
    except:
        df = pd.DataFrame()

    conn.close()
    return df

def salvar_pagina_cache(df, chave):
    if df.empty:
        return

    df = df.copy()
    df["cache_key"] = chave

    conn = sqlite3.connect(DB_PATH)
    df.to_sql("cache_paginas", conn, if_exists="append", index=False)
    conn.close()

# -------------------------------
# 🔹 SALVAR NO BANCO
# -------------------------------
def salvar_no_banco(df, tabela):
    if df is None or df.empty:
        return

    df = df.copy()

    # 🔥 REMOVE DICT/LIST (CONVERTE PRA STRING)
    for col in df.columns:
        df[col] = df[col].apply(
            lambda x: json.dumps(x, ensure_ascii=False)
            if isinstance(x, (dict, list))
            else x
        )

    conn = sqlite3.connect(DB_PATH)
    df.to_sql(tabela, conn, if_exists="append", index=False)
    conn.close()
# -------------------------------salvar_pagina_cache
# 🔹 CARREGAR DO BANCO
# -------------------------------
def carregar_do_banco(tabela):
    if not os.path.exists(DB_PATH):
        return pd.DataFrame()

    conn = sqlite3.connect(DB_PATH)

    try:
        df = pd.read_sql(
            f"SELECT * FROM '{tabela}'",
            conn
        )
    except:
        df = pd.DataFrame()

    conn.close()
    return df

# -------------------------------ultimo_envio
# 🔹 CONTROLE DE EXPIRAÇÃO
# -------------------------------
def banco_expirado(minutos=30):
    if not os.path.exists(DB_PATH):
        return True

    ultima_mod = os.path.getmtime(DB_PATH)
    agora = time.time()

    return (agora - ultima_mod) > (minutos * 60)


# ------------------------------- df_cache
# 🔹 FUNÇÃO PRINCIPAL (SMART LOAD)
# -------------------------------
def carregar_ou_buscar(nome_tabela, func_busca, *args):
    """
    Cache inteligente baseado nos parâmetros da consulta
    """
    chave = f"{nome_tabela}_{hash(str(args))}"
    # 🔑 CRIA CHAVE ÚNICA BASEADA NOS PARÂMETROS
    def gerar_chave(nome, pagina, parametros):
        return f"{nome}_p{pagina}_{hash(str(parametros))}"

    def carregar_pagina_cache(nome, pagina, *args):
        chave = gerar_chave(nome, pagina, *args)
        return carregar_do_banco(chave)

    def salvar_pagina_cache(df, nome, pagina, *args):
        chave = gerar_chave(nome, pagina, *args)
        salvar_no_banco(df, chave)

    if banco_expirado(30):
        df = func_busca(*args)
        salvar_no_banco(df, chave)
        return df

    df = carregar_do_banco(chave)

    if df.empty:
        df = func_busca(*args)
        salvar_no_banco(df, chave)

    return df

st.set_page_config(layout="wide")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ARQUIVO_CONTROLE = "ultimo_envio.txt"

def ja_enviou_hoje():

    if not os.path.exists(ARQUIVO_CONTROLE):
        with open(ARQUIVO_CONTROLE, "w") as f:
            f.write("")

    if not os.path.exists(ARQUIVO_CONTROLE):
        return False

    with open(ARQUIVO_CONTROLE, "r") as f:
        ultima_data = f.read().strip()

    hoje = datetime.now().strftime("%Y-%m-%d")

    return ultima_data == hoje


def salvar_envio():
    hoje = datetime.now().strftime("%Y-%m-%d")

    with open(ARQUIVO_CONTROLE, "w") as f:
        f.write(hoje)
# ============================== ImageFont
# CONFIG API 
# ==============================
BASE_URL = "https://apiintegracao.milvus.com.br/api/chamado/listagem"
TOKEN = "zayqrw9btEIxTtM26w0JjCWrW8GS4QsVXFkAePTm2unCU6DsTrmWUD7lOP94lxocEgXRZBdlxTmSPrkjM9vdCrFEytnsbZOQFxgK7"

headers = {
    'Content-Type': 'application/json',
    'Authorization': TOKEN
}

# ==============================
# CONFIG DISCORD
# ==============================value
WEBHOOK_URL = "https://discord.com/api/webhooks/1480567192859705537/jMFT7DKm2ShNEjXSZCeULjgMpq-owh__PNIUSjR9Ezpfeh2SZEDfc-wwzyqQG1RIlfCa"

# ==============================chave 
# ⚡ CONTROLE DE PAGINAÇÃO
# ==============================

st.sidebar.subheader("⚡ Performance da Consulta")

qtd_paginas = st.sidebar.number_input(
    "Quantidade de páginas da API",
    min_value=1,
    max_value=10,    
    value=1,  # 🔥 padrão
    step=1
    
)

st.sidebar.caption("⚠️ Mais páginas = mais lento")
@st.cache_data(ttl=120)
def buscar_dados(qtd_paginas):

    st.write(f"🔍 DEBUG → Buscando {qtd_paginas} páginas")

    todos = []

    hoje = datetime.now()
    data_inicio_api = (hoje - timedelta(days=60)).strftime("%Y-%m-%d 00:00:00")
    data_fim_api = hoje.strftime("%Y-%m-%d 23:59:59")

    progress = st.progress(0, text="Buscando dados...")

    for pagina in range(1, qtd_paginas + 1):

        url = f"{BASE_URL}?total_registros=1000&pagina={pagina}"

        payload = json.dumps({
            "filtro_body": {
                "data_criacao_inicio": data_inicio_api,
                "data_criacao_fim": data_fim_api
            }
        })

        try:
            response = requests.post(
                url,
                headers=headers,
                data=payload,
                timeout=(5, 30)  # conexão, leitura
            )

            if response.status_code != 200:
                st.error(f"Erro API: {response.status_code}")
                break

            data = response.json()
            lista = data.get("lista", [])

            if not lista:
                break

            todos.extend(lista)

            progress.progress(
                pagina / qtd_paginas,
                text=f"Página {pagina} de {qtd_paginas}"
            )

            # 🔥 para antes se acabar os dados
            if len(lista) < 1000:
                break

        except Exception as e:
            st.error(f"Erro: {e}")
            break

    progress.empty()

    df = pd.DataFrame(todos)

    if df.empty:
        return df

    # =========================
    # NORMALIZAÇÃO
    # =========================
    for col in ["categoria_primaria", "categoria_secundaria", "mesa_trabalho"]:
        if col in df.columns:
            df[col] = df[col].apply(
                lambda x: x.get("text") if isinstance(x, dict) else x
            )

    if "data_criacao" in df.columns:
        df["data_criacao"] = pd.to_datetime(df["data_criacao"], errors="coerce")

    return df

df = carregar_ou_buscar("tickets", buscar_dados, qtd_paginas)


@st.cache_data(ttl=600)
def buscar_whatsapp(qtd_paginas, data_inicio, data_fim):

    BASE_URL_CHAT = "https://apiintegracao.milvus.com.br/api/chat/listagem"

    todos = []
    pagina = 1

    paginas_sem_resultado = 0
    LIMITE_PAGINAS_VAZIAS = 10
    MAX_RETRIES = 3

    payload = json.dumps({
        "filtro_body": {
            "data_inicio_atendimento_inicio": data_inicio.strftime("%Y-%m-%d %H:%M:%S"),
            "data_inicio_atendimento_fim": data_fim.strftime("%Y-%m-%d %H:%M:%S")
        }
    })

    progress = st.progress(0, text="Buscando WhatsApp...")

    while True:

        url = f"{BASE_URL_CHAT}?total_registros=1000&pagina={pagina}"

        # =========================================
        # 🚀 CACHE POR PÁGINA
        # =========================================tabela
        chave = carregar_pagina_cache(
            "whatsapp",
            pagina,
            (data_inicio, data_fim)
        )
        df_cache = carregar_pagina_cache(chave)

        if not df_cache.empty:
            df_page = df_cache.copy()
            origem = "⚡ Cache"

        else:
            origem = "📡 API"

            retry = 0
            response = None

            # 🔁 RETRY AUTOMÁTICO
            while retry < MAX_RETRIES:
                try:
                    response = requests.post(
                        url,
                        headers=headers,
                        data=payload,
                        timeout=40
                    )

                    if response.status_code == 200:
                        break

                except:
                    pass

                retry += 1
                time.sleep(1)

            # ❌ ERRO → PARA (evita loop infinito)
            if response is None or response.status_code != 200:
                st.warning(f"⚠️ Falha na página {pagina}, abortando busca")
                break

            data = response.json()
            lista = data.get("lista", [])

            if not lista:
                break

            df_page = pd.DataFrame(lista)

            # 💾 SALVA CACHE DA PÁGINA
            salvar_pagina_cache(df_page, chave)

        # =========================================
        # 🔄 PROCESSAMENTO
        # =========================================
        if "data_inicio_atendimento" not in df_page.columns:
            pagina += 1
            continue

        df_page["data_inicio_atendimento"] = pd.to_datetime(
            df_page["data_inicio_atendimento"],
            errors="coerce"
        )

        df_page = df_page[df_page["data_inicio_atendimento"].notna()]

        df_filtrado = df_page[
            (df_page["data_inicio_atendimento"] >= data_inicio) &
            (df_page["data_inicio_atendimento"] <= data_fim)
        ]

        todos.extend(df_filtrado.to_dict("records"))

        # =========================================
        # 🧠 PARADAS INTELIGENTES
        # =========================================

        # se já passou do período → para
        if not df_page.empty and df_page["data_inicio_atendimento"].max() < data_inicio:
            break

        # controle de páginas vazias
        if len(df_filtrado) == 0:
            paginas_sem_resultado += 1
        else:
            paginas_sem_resultado = 0

        if paginas_sem_resultado >= LIMITE_PAGINAS_VAZIAS:
            break

        # progresso
        progress.progress(
            min(pagina / qtd_paginas, 1.0),
            text=f"{origem} página {pagina}"
        )

        pagina += 1

    progress.empty()

    # =========================================
    # 📦 DATAFRAME FINAL
    # =========================================
    df_chat = pd.DataFrame(todos)

    if df_chat.empty:
        return df_chat

    # ==============================
    # 🔥 TRATAMENTO FINAL
    # ==============================

    df_chat["data_inicio_atendimento"] = pd.to_datetime(
        df_chat.get("data_inicio_atendimento"),
        errors="coerce"
    )

    if "data_mensagem" in df_chat.columns:
        df_chat["data_mensagem"] = pd.to_datetime(
            df_chat["data_mensagem"],
            errors="coerce"
        )

    # técnico pode vir como dict
    if "tecnico_ativo" in df_chat.columns:
        df_chat["tecnico_ativo"] = df_chat["tecnico_ativo"].apply(
            lambda x: x.get("text") if isinstance(x, dict) else x
        )

    # remove duplicados
    if "id" in df_chat.columns:
        df_chat = df_chat.drop_duplicates(subset=["id"])

    return df_chat
# ==============================data_inicio 
# VALIDAÇÕES
# ==============================


if df is None or df.empty:
    st.error("❌ Nenhum dado retornado")
    st.stop()

if "data_criacao" not in df.columns:
    st.error("❌ Coluna data_criacao não existe")
    st.write(df.columns.tolist())  # DEBUG
    st.stop()

df = df.dropna(subset=["data_criacao"])

if df.empty:
    st.error("❌ Datas inválidas")
    st.stop()

# ============================== cache_data
# FILTRO DE MESA DE TRABALHO (NOVO)
# ==============================
st.sidebar.subheader("Mesa de Trabalho")

opcoes_mesa = sorted(df["mesa_trabalho"].dropna().unique())

mesas_selecionadas = st.sidebar.multiselect(
    "Selecione as mesas",
    opcoes_mesa,
    default=["Suporte Técnico"]
)

if "mesa_trabalho" in df.columns:
    df = df[df["mesa_trabalho"].isin(mesas_selecionadas)]

if df.empty:
    st.warning("⚠️ Nenhum dado para as mesas selecionadas")
    st.stop()

# ==============================data_inicio 
# 🚀 FILTRO AVANÇADO DE DATA
# ==============================
st.sidebar.subheader("📅 Filtro de Data Avançado")

modo_data = st.sidebar.radio(
    "Tipo de filtro",
    [
        "Período",
        "Dias específicos",
        "Dias da semana",
        "Comparar períodos"
    ]
)

# ------------------------------ qtd_paginas 
# 1️⃣ PERÍODO
# ------------------------------
if modo_data == "Período":
    data_inicio, data_fim = st.sidebar.date_input(
        "Período",
        [
            df["data_criacao"].min().date(),
            df["data_criacao"].max().date()
        ]
    )

    df = df[
        (df["data_criacao"].dt.date >= data_inicio) &
        (df["data_criacao"].dt.date <= data_fim)
    ]

# ------------------------------data_inicio
# 2️⃣ DIAS ESPECÍFICOS
# ------------------------------
elif modo_data == "Dias específicos":
    datas_disponiveis = sorted(df["data_criacao"].dt.date.dropna().unique())

    datas_selecionadas = st.sidebar.multiselect(
        "Selecione os dias",
        datas_disponiveis,
        default=datas_disponiveis[-1:]
    )

    if datas_selecionadas:
        df = df[df["data_criacao"].dt.date.isin(datas_selecionadas)]
    else:
        st.warning("⚠️ Selecione ao menos um dia")
        st.stop()

    data_inicio = min(datas_selecionadas)
    data_fim = max(datas_selecionadas)

# ------------------------------
# 3️⃣ DIAS DA SEMANA
# ------------------------------
elif modo_data == "Dias da semana":

    dias_map = {
        "Segunda": 0,
        "Terça": 1,
        "Quarta": 2,
        "Quinta": 3,
        "Sexta": 4,
        "Sábado": 5,
        "Domingo": 6
    }

    dias_escolhidos = st.sidebar.multiselect(
        "Selecione os dias da semana",
        list(dias_map.keys()),
        default=["Segunda", "Terça", "Quarta", "Quinta", "Sexta"]
    )

    if dias_escolhidos:
        numeros = [dias_map[d] for d in dias_escolhidos]

        df = df[df["data_criacao"].dt.weekday.isin(numeros)]
    else:
        st.warning("⚠️ Selecione ao menos um dia")
        st.stop()

    data_inicio = df["data_criacao"].min().date()
    data_fim = df["data_criacao"].max().date()

# ------------------------------
# 4️⃣ COMPARAR PERÍODOS
# ------------------------------
else:
    st.sidebar.markdown("### 🔁 Período 1")

    p1_inicio, p1_fim = st.sidebar.date_input(
        "Período 1",
        [
            df["data_criacao"].min().date(),
            df["data_criacao"].max().date()
        ],
        key="p1"
    )

    st.sidebar.markdown("### 🔁 Período 2")

    p2_inicio, p2_fim = st.sidebar.date_input(
        "Período 2",
        [
            df["data_criacao"].min().date(),
            df["data_criacao"].max().date()
        ],
        key="p2"
    )

    df_p1 = df[
        (df["data_criacao"].dt.date >= p1_inicio) &
        (df["data_criacao"].dt.date <= p1_fim)
    ]

    df_p2 = df[
        (df["data_criacao"].dt.date >= p2_inicio) &
        (df["data_criacao"].dt.date <= p2_fim)
    ]

    # KPI comparativo
    st.markdown("## 📊 Comparação de Períodos")

    colA, colB = st.columns(2)

    colA.metric("Período 1", len(df_p1))
    colB.metric("Período 2", len(df_p2))

    # mantém df principal como período 1 (para não quebrar resto)
    df = df_p1

    data_inicio = p1_inicio
    data_fim = p1_fim

# ==============================
# 📲 BUSCAR WHATSAPP (COM CACHE LOCAL)
# ==============================

data_inicio = pd.to_datetime(data_inicio)
data_fim = pd.to_datetime(data_fim) + pd.Timedelta(days=1)

df_chat = carregar_ou_buscar(
    "whatsapp",
    buscar_whatsapp,
    qtd_paginas,
    data_inicio,
    data_fim
)

# segurança extra (evita crash)
if df_chat is None:
    df_chat = pd.DataFrame()

# segue o fluxo normal
total_tickets = len(df)

# ============================== cache_data
# AJUSTES
# ==============================
def safe_col(nome):
    return nome if nome in df.columns else None

col_cliente = safe_col("cliente")
col_tecnico = safe_col("tecnico")
col_cat_prim = safe_col("categoria_primaria")
col_cat_sec = safe_col("categoria_secundaria")

if col_tecnico:
    df["tecnico"] = df["tecnico"].astype(str).str.split().str[0]

# ==============================
# FUNÇÃO TOP
# ==============================
def top(df, coluna, n=None):
    if not coluna or coluna not in df.columns:
        return pd.DataFrame()

    df_count = df[coluna].value_counts()

    if n:
        df_count = df_count.head(n)

    df_top = df_count.reset_index()
    df_top.columns = [coluna, "Quantidade"]

    return df_top

top_cat_prim = top(df, col_cat_prim, 5)
top_cat_sec = top(df, col_cat_sec, 5)
top_clientes = top(df, col_cliente, 10)
top_tecnicos = top(df, col_tecnico)

# ==============================
# PALETA DE CORES
# ==============================
cores = [
    "#4E79A7",  # azul
    "#F28E2B",  # laranja
    "#59A14F",  # verde
    "#E15759",  # vermelho
    "#B07AA1"   # roxo
]

# ==============================
# DASHBOARD
# ==============================
st.title(f"📊 Painel NOC {' - '.join(mesas_selecionadas)}")
# 👇 NOVO KPI
st.metric("🎫 Total de Tickets", total_tickets)

col1, col2 = st.columns(2)

fig1 = fig2 = fig3 = fig4 = None

if not top_cat_prim.empty:
    fig1 = px.bar(
    top_cat_prim,
    x=col_cat_prim,
    y="Quantidade",
    title="Top Categoria Primária",
    text="Quantidade",
    color=col_cat_prim,
    color_discrete_sequence=cores

    )
    fig1.update_traces(textposition='outside')
    fig1.update_layout(uniformtext_minsize=10, uniformtext_mode='hide', showlegend=False)
    col1.plotly_chart(fig1, use_container_width=True)

if not top_cat_sec.empty:
    fig2 = px.bar(
    top_cat_sec,
    x=col_cat_sec,
    y="Quantidade",
    title="Top Categoria Secundária",
    text="Quantidade",
    color=col_cat_sec,
    color_discrete_sequence=cores

    )
    fig2.update_traces(textposition='outside')
    fig2.update_layout(uniformtext_minsize=10, uniformtext_mode='hide', showlegend=False)
    col2.plotly_chart(fig2, use_container_width=True)

col3, col4 = st.columns(2)

if not top_clientes.empty:
    fig3 = px.bar(
        top_clientes,
        x=col_cliente,
        y="Quantidade",
        title="Top Clientes",
        text="Quantidade",
        color=col_cliente,
        color_discrete_sequence=cores
    )
    fig3.update_traces(textposition='outside')
    fig3.update_layout(
        xaxis_tickangle=-45,
        uniformtext_minsize=10,
        uniformtext_mode='hide', showlegend=False
    )
    col3.plotly_chart(fig3, use_container_width=True)

if not top_tecnicos.empty:
    fig4 = px.bar(
        top_tecnicos,
        x=col_tecnico,
        y="Quantidade",
        title="Tickets por Técnico",
        text="Quantidade",
        color=col_tecnico,
        color_discrete_sequence=cores
    )
    fig4.update_traces(textposition='outside')
    fig4.update_layout(
        uniformtext_minsize=10,
        uniformtext_mode='hide',
        height=700,
        showlegend=False
    )
    col4.plotly_chart(fig4, use_container_width=True)

# ============================== data_inicio_atendimento
# 📲 RELATÓRIO WHATSAPP
# ==============================
st.markdown("## 📲 Atendimentos WhatsApp")

if df_chat.empty:
    st.warning("⚠️ Nenhum atendimento encontrado")

else:
    df_wpp = df_chat.copy()

    # apenas finalizados
    df_wpp = df_wpp[df_wpp["status"] == "Finalizado"]

    # KPI
    total_wpp = len(df_wpp)
    st.metric("📲 Total Finalizados WhatsApp", total_wpp)

    if "tecnico_ativo" in df_wpp.columns:

        resumo_wpp = (
            df_wpp[df_wpp["tecnico_ativo"].notna()]
            .groupby("tecnico_ativo")
            .size()
            .reset_index(name="Atendimentos Finalizados")
            .sort_values(by="Atendimentos Finalizados", ascending=False)
        )

        fig_wpp = px.bar(
            resumo_wpp,
            x="tecnico_ativo",
            y="Atendimentos Finalizados",
            text="Atendimentos Finalizados",
            color="tecnico_ativo",
            color_discrete_sequence=cores
        )

        fig_wpp.update_traces(textposition='outside')
        fig_wpp.update_layout(showlegend=False)

        st.plotly_chart(fig_wpp, use_container_width=True)
# ==============================
# 📊 COMPARATIVO ENTRE MESAS
# ==============================
st.markdown("## 📊 Comparativo entre Mesas")

if "mesa_trabalho" in df.columns and col_cat_prim:
    df_cat_mesa = df.groupby(
        ["categoria_primaria", "mesa_trabalho"]
    ).size().reset_index(name="Quantidade")

    fig_comp = px.bar(
        df_cat_mesa,
        x="categoria_primaria",
        y="Quantidade",
        color="mesa_trabalho",
        barmode="group",  # pode trocar por "stack"
        title="Categorias por Mesa",
        text="Quantidade"
    )

    fig_comp.update_traces(textposition='outside')
    fig_comp.update_layout(xaxis_tickangle=-30)

    st.plotly_chart(fig_comp, use_container_width=True)

# ==============================
# 📊 GRÁFICOS POR MESA
# ==============================
st.markdown("## 📊 Análise por Mesa")

mesas = df["mesa_trabalho"].dropna().unique()

for mesa in mesas:
    
    st.markdown(f"### 🧩 {mesa}")

    df_mesa = df[df["mesa_trabalho"] == mesa]

    colA, colB = st.columns(2)

    # Top Categorias
    if col_cat_prim:
        top_cat = df_mesa[col_cat_prim].value_counts().head(5).reset_index()
        top_cat.columns = [col_cat_prim, "Quantidade"]

        fig_cat = px.bar(
            top_cat,
            x=col_cat_prim,
            y="Quantidade",
            title=f"Top Categorias - {mesa}",
            text="Quantidade"
        )

        fig_cat.update_traces(textposition='outside')
        fig_cat.update_layout(xaxis_tickangle=-30)

        colA.plotly_chart(fig_cat, use_container_width=True)

    # Top Técnicos
    if col_tecnico:
        top_tec = df_mesa[col_tecnico].value_counts().reset_index()
        top_tec.columns = [col_tecnico, "Quantidade"]

        fig_tec = px.bar(
            top_tec,
            x=col_tecnico,
            y="Quantidade",
            title=f"Técnicos - {mesa}",
            text="Quantidade"
        )

        fig_tec.update_traces(textposition='outside')

        colB.plotly_chart(fig_tec, use_container_width=True)


# ==============================
# ▶️ CHAMADOS EM ATENDIMENTO (PLAY)
# ==============================

import streamlit as st
import pandas as pd
import re
import html
from datetime import datetime
from streamlit_autorefresh import st_autorefresh

# 🔄 AUTO REFRESH
st_autorefresh(interval=300000, key="refresh")

# ==============================max_value
# 🔹 FUNÇÕES
# ==============================
def limpar_html(texto):
    if pd.isna(texto):
        return "Sem assunto"
    texto = html.unescape(str(texto))
    texto = texto.replace("<br>", " ").replace("<br/>", " ")
    texto = re.sub(r"<.*?>", "", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto

def limitar_texto(texto, limite=80):
    return texto if len(texto) <= limite else texto[:limite] + "..."

# ==============================
# 🔥 FILTRO ATENDENDO
# ==============================
df_play = df[
    df["status"].astype(str).str.lower().str.contains("atend")
].copy()

st.markdown("## ▶️ Chamados em Atendimento (PLAY)")
st.markdown(f"### 📊 Total em atendimento: **{len(df_play)}**")

if df_play.empty:
    st.success("✅ Nenhum chamado em atendimento")

else:
    df_play["assunto"] = df_play["assunto"].apply(limpar_html)
    df_play["assunto"] = df_play["assunto"].apply(limitar_texto)

    df_play["inicio_play"] = pd.to_datetime(df_play["data_modificacao"], errors="coerce")

    agora = datetime.now()
    df_play["Tempo (min)"] = (
        (agora - df_play["inicio_play"]).dt.total_seconds() / 60
    ).fillna(0).astype(int)

    df_play = df_play.sort_values("Tempo (min)", ascending=False)

    # ==============================
    # 🎯 CARDS
    # ==============================
    for i, row in df_play.iterrows():

        tempo = row["Tempo (min)"]

        if tempo >= 120:
            cor = "#ff0000"   # vermelho forte
            alerta = "🚨 CRÍTICO"

        elif tempo >= 60:
            cor = "#ff9900"   # laranja forte
            alerta = "⚠️ ATRASADO"

        elif tempo >= 30:
            cor = "#ffd700"   # amarelo
            alerta = "⏳ ATENÇÃO"

        elif tempo >= 10:
            cor = "#00cc66"   # verde
            alerta = "🟢 EM ANDAMENTO"

        else:
            cor = "#3399ff"   # azul (recém iniciado)
            alerta = "🔵 RECENTE"

        tecnico = html.escape(str(row["tecnico"]))
        cliente = html.escape(str(row["cliente"]))
        assunto = html.escape(str(row["assunto"]))

        link_ticket = f"https://app.milvus.com.br/tickets/details/{row['id']}"

        # 🔹 CARD VISUAL
        st.markdown(f"""
        <div style="
            background:#111;
            padding:12px;
            border-radius:10px;
            margin-bottom:5px;
            border-left:6px solid {cor};
            color:white;
        ">
            <b>👨‍🔧 {tecnico}</b> |
            🏢 {cliente} <br>

            🧩 {assunto} <br>

            <span style="color:{cor}; font-weight:bold;">
                {alerta}
            </span><br>

            ⏱️ {tempo} min |
            🕒 {row['inicio_play'].strftime("%d/%m %H:%M")}
        </div>
        """, unsafe_allow_html=True)

        # 🔗 BOTÃO REAL (FUNCIONA SEM BUG)
        if st.button(f"🔗 Abrir Ticket {row['id']}", key=f"btn_{i}"):
            st.markdown(f"[Abrir no sistema]({link_ticket})")
        
# ==============================
# GERAR IMAGEM ÚNICA
# ==============================
def gerar_imagem_unica(figs, data_inicio, data_fim, total_tickets):
    imagens = []

    for i, fig in enumerate(figs):
        if fig is None:
            continue

        nome = f"temp_{i}.png"
        fig.write_image(nome, width=800, height=500)
        imagens.append(Image.open(nome))

    if not imagens:
        return None

    imagem_final = Image.new("RGB", (1600, 1100), "white")

    posicoes = [(0,100), (800,100), (0,600), (800,600)]

    for img, pos in zip(imagens, posicoes):
        imagem_final.paste(img.resize((800, 500)), pos)

    draw = ImageDraw.Draw(imagem_final)

    texto = f"{' | '.join(mesas_selecionadas)} | {data_inicio} até {data_fim} | Total: {total_tickets}"

    try:
        fonte = ImageFont.truetype("arial.ttf", 40)
    except:
        
        fonte = ImageFont.load_default()

    draw.text((20, 30), texto, fill="black", font=fonte)

    caminho = "dashboard_final.png"
    imagem_final.save(caminho)

    for i in range(len(imagens)):
        os.remove(f"temp_{i}.png")

    return caminho

# ==============================cache_milvus
# ENVIAR DISCORD
# ==============================

def enviar_imagem_unica(caminho):
    if not caminho:
        return

    with open(caminho, "rb") as f:
        requests.post(WEBHOOK_URL, files={"file": f})

def ja_enviou_hoje():
    if not os.path.exists(ARQUIVO_CONTROLE):
        return False

    with open(ARQUIVO_CONTROLE, "r") as f:
        ultima_data = f.read().strip()

    hoje = datetime.now().strftime("%Y-%m-%d")

    return ultima_data == hoje

# ==============================
# 🤖 ENVIO AUTOMÁTICO
# ==============================
hora_envio = 8
agora = datetime.now()

if agora.hour >= hora_envio and not ja_enviou_hoje():

    # =========================
    # SALVA DF ORIGINAL
    # =========================
    df_backup = df.copy()

    # =========================
    # FILTRO AUTOMÁTICO
    # =========================
    ontem = datetime.now() - timedelta(days=1)

    df = df_backup[
        (df_backup["data_criacao"].dt.date == ontem.date()) &
        (df_backup["mesa_trabalho"] == "Suporte Técnico")
    ]

    if df.empty:
        print("Sem dados para envio automático")

    else:
        # =========================
        # RECALCULA TOPS (IMPORTANTE)
        # =========================
        top_cat_prim = top(df, col_cat_prim, 5)
        top_cat_sec = top(df, col_cat_sec, 5)
        top_clientes = top(df, col_cliente, 10)
        top_tecnicos = top(df, col_tecnico)

        # =========================
        # RECRIA OS MESMOS GRÁFICOS st
        # =========================
        fig1 = fig2 = fig3 = fig4 = None

        if not top_cat_prim.empty:
            fig1 = px.bar(
            top_cat_prim,
            x=col_cat_prim,
            y="Quantidade",
            title="Top Categoria Primária",
            text="Quantidade",
            color=col_cat_prim,
            color_discrete_sequence=cores

            )
        fig1.update_traces(textposition='outside')
        fig1.update_layout(uniformtext_minsize=10, uniformtext_mode='hide', showlegend=False)
        col1.plotly_chart(fig1, use_container_width=True)

        if not top_cat_sec.empty:
            fig2 = px.bar(
            top_cat_sec,
            x=col_cat_sec,
            y="Quantidade",
            title="Top Categoria Secundária",
            text="Quantidade",
            color=col_cat_sec,
            color_discrete_sequence=cores

            )
        fig2.update_traces(textposition='outside')
        fig2.update_layout(uniformtext_minsize=10, uniformtext_mode='hide', showlegend=False)
        col2.plotly_chart(fig2, use_container_width=True, key="top5_categorias_secundaria")

        col3, col4 = st.columns(2)

    if not top_clientes.empty:
        fig3 = px.bar(
        top_clientes,
        x=col_cliente,
        y="Quantidade",
        title="Top Clientes",
        text="Quantidade",
        color=col_cliente,
        color_discrete_sequence=cores
        )
    fig3.update_traces(textposition='outside')
    fig3.update_layout(
        xaxis_tickangle=-45,
        uniformtext_minsize=10,
        uniformtext_mode='hide', showlegend=False
        )
    col3.plotly_chart(fig3, use_container_width=True, key="top5_categorias")

    if not top_tecnicos.empty:
        fig4 = px.bar(
        top_tecnicos,
        x=col_tecnico,
        y="Quantidade",
        title="Tickets por Técnico",
        text="Quantidade",
        color=col_tecnico,
        color_discrete_sequence=cores
    )
        fig4.update_traces(textposition='outside')
        fig4.update_layout(
        uniformtext_minsize=10,
        uniformtext_mode='hide',
        height=700,
        showlegend=False
    )
        col4.plotly_chart(fig4, use_container_width=True, key="top5_clientes")

        # =========================
        # ENVIO
        # =========================
        caminho = gerar_imagem_unica(
            [fig1, fig2, fig3, fig4],
            ontem.date(),
            ontem.date(),
            len(df)
        )

        enviar_imagem_unica(caminho)
        salvar_envio()

        print("Relatório automático enviado!")

    # =========================
    # RESTAURA DF ORIGINAL
    # =========================
    df = df_backup
# ==============================total_tickets 
# BOTÃO
# ==============================
st.markdown("### 🚀 Enviar painel completo para Discord")

if st.button("Enviar agora"):
    with st.spinner("Gerando imagem..."):
        caminho = gerar_imagem_unica(
            [fig1, fig2, fig3, fig4],
            data_inicio,
            data_fim,
            total_tickets
        )

        enviar_imagem_unica(caminho)

    st.success("✅ Painel enviado em imagem única!")

#========================================================
#excel
#========================================================

# ========================================================
# ?? EXPORTAÇÃO EXCEL (VERSÃO FINAL ESTÁVEL)
# ========================================================
from io import BytesIO
import pandas as pd
import json
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
import streamlit as st

# ========================================================
# ?? FUNÇÃO LIMPEZA (ANTI-EXCEL CORROMPIDO)
# ========================================================
def limpar_df_para_excel(df):
    df = df.copy()

    # remove timezone
    for col in df.columns:
        if "data" in col:
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.tz_localize(None)

    # dict/list ? string
    for col in df.columns:
        df[col] = df[col].apply(
            lambda x: json.dumps(x, ensure_ascii=False)
            if isinstance(x, (dict, list)) else x
        )

    # remove caracteres ilegais
    def limpar_texto(valor):
        if isinstance(valor, str):
            return re.sub(r'[\x00-\x1F]+', '', valor)
        return valor

    df = df.applymap(limpar_texto)

    # limita tamanho
    def limitar(valor):
        if isinstance(valor, str):
            return valor[:30000]
        return valor

    df = df.applymap(limitar)

    return df


# ========================================================
# ?? EXCEL PRINCIPAL
# ========================================================
def gerar_excel(df, df_chat, mesas_selecionadas):
    output = BytesIO()

    if df is None or df.empty:
        return output

    df_excel = df.copy()

    df_excel["data_criacao"] = pd.to_datetime(df_excel["data_criacao"], errors="coerce")
    df_excel = df_excel.dropna(subset=["data_criacao"])

    if mesas_selecionadas:
        df_excel = df_excel[df_excel["mesa_trabalho"].isin(mesas_selecionadas)]

    if df_excel.empty:
        return output

    df_excel["Mes"] = df_excel["data_criacao"].dt.strftime("%m-%Y")
    df_excel["Data"] = df_excel["data_criacao"].dt.strftime("%d/%m")
    df_excel["Dia"] = df_excel["data_criacao"].dt.day_name()

    mapa_dias = {
        "Monday": "Segunda", "Tuesday": "Terça", "Wednesday": "Quarta",
        "Thursday": "Quinta", "Friday": "Sexta",
        "Saturday": "Sábado", "Sunday": "Domingo"
    }

    df_excel["Dia"] = df_excel["Dia"].map(mapa_dias)

    agrupado = (
        df_excel
        .groupby(["Mes", "Data", "Dia", "mesa_trabalho"])
        .size()
        .reset_index(name="Qtd")
    )

    linhas = []

    for (mes, data, dia), grupo in agrupado.groupby(["Mes", "Data", "Dia"]):
        linha = {"Dia": dia, "Data": data}
        total_dia = 0

        for mesa in mesas_selecionadas:
            valor = grupo.loc[grupo["mesa_trabalho"] == mesa, "Qtd"].sum()
            linha[mesa] = int(valor)
            total_dia += valor

        linha["Total Dia"] = int(total_dia)
        linhas.append(linha)

    df_final = pd.DataFrame(linhas)

    df_final["Data_ordem"] = pd.to_datetime(
        df_final["Data"] + f"/{datetime.now().year}",
        format="%d/%m/%Y",
        errors="coerce"
    )

    df_final = df_final.sort_values("Data_ordem").drop(columns="Data_ordem")

    total = {"Dia": "TOTAL", "Data": ""}
    for mesa in mesas_selecionadas:
        total[mesa] = int(df_final[mesa].sum())

    total["Total Dia"] = int(df_final["Total Dia"].sum())
    df_final = pd.concat([df_final, pd.DataFrame([total])], ignore_index=True)

    ranking = (
        df_excel.groupby("mesa_trabalho")
        .size()
        .reset_index(name="Total Tickets")
        .sort_values("Total Tickets", ascending=False)
    )

    top_dias = (
        df_final[df_final["Dia"] != "TOTAL"]
        .sort_values("Total Dia", ascending=False)
        .head(10)
    )

    # =========================
    # ?? WHATSAPP
    # ========================= Baixar
    df_wpp_excel = pd.DataFrame()
    resumo_wpp = pd.DataFrame()

    if df_chat is not None and not df_chat.empty:
        df_wpp_excel = df_chat.copy()

        if "status" in df_wpp_excel.columns:
            df_wpp_excel = df_wpp_excel[df_wpp_excel["status"] == "Finalizado"]

        if "tecnico_ativo" in df_wpp_excel.columns:
            resumo_wpp = (
                df_wpp_excel[df_wpp_excel["tecnico_ativo"].notna()]
                .groupby("tecnico_ativo")
                .size()
                .reset_index(name="Atendimentos Finalizados")
                .sort_values(by="Atendimentos Finalizados", ascending=False)
            )

        df_wpp_excel = limpar_df_para_excel(df_wpp_excel)

    # =========================
    # ?? SALVAR
    # =========================
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_final.to_excel(writer, sheet_name="Geral", index=False)
        ranking.to_excel(writer, sheet_name="Ranking", index=False)
        top_dias.to_excel(writer, sheet_name="Top Dias", index=False)

        if not df_wpp_excel.empty:
            df_wpp_excel.to_excel(writer, sheet_name="WhatsApp - Dados", index=False)

        if not resumo_wpp.empty:
            resumo_wpp.to_excel(writer, sheet_name="WhatsApp - Resumo", index=False)

    output.seek(0)
    wb = load_workbook(output)

    # =========================carregar_ou_buscar
    # ?? FORMATAÇÃO
    # =========================
    ws = wb["Geral"]
    max_row = ws.max_row
    max_col = ws.max_column

    tabela = Table(ref=f"A1:{chr(64+max_col)}{max_row}", displayName="TabelaGeral")
    tabela.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tabela)

    ws.conditional_formatting.add(
        f"C2:{chr(64+max_col)}{max_row-1}",
        ColorScaleRule(
            start_type='min', start_color='63BE7B',
            mid_type='percentile', mid_value=50, mid_color='FFEB84',
            end_type='max', end_color='F8696B'
        )
    )

    chart = BarChart()
    data = Reference(ws, min_col=3, max_col=max_col, min_row=1, max_row=max_row-1)
    cats = Reference(ws, min_col=2, min_row=2, max_row=max_row-1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.title = "Tickets por Dia"
    ws.add_chart(chart, "H2")

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return final_output


# ========================================================carregar_do_banco
# ?? BOTÕES STREAMLIT
# ========================================================
excel_file = gerar_excel(df, df_chat, mesas_selecionadas)

st.download_button(
    label="?? Baixar Excel Geral",
    data=excel_file,
    file_name="relatorio_geral.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# ========================================================data_inicio_atendimento_inicio
# ?? ATUALIZAR DADOS
# ========================================================
if st.sidebar.button("🔄 Atualizar dados agora"):

    # 🔥 REMOVE BANCO ANTIGO (FORÇA ATUALIZAÇÃO REAL)
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)

    df = buscar_dados(qtd_paginas)
    salvar_no_banco(df, "tickets")

    df_chat = buscar_whatsapp(qtd_paginas, data_inicio, data_fim)
    salvar_no_banco(df_chat, f"whatsapp_{hash(str((qtd_paginas, data_inicio, data_fim)))}")

    st.success("✅ Dados atualizados com sucesso!")